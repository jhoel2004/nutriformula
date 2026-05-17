# ui/tab_ingredientes.py
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
                             QTableWidget, QTableWidgetItem, QLineEdit,
                             QComboBox, QHeaderView, QMessageBox, QFileDialog,
                             QAbstractItemView)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QBrush

import openpyxl

from app.database import get_all_insumos, insert_insumo, update_insumo, delete_insumo
from app.calculator import auditar_insumo
from ui.dialogs import InsumoDialog

HEADERS = ["Nombre", "Proteína%", "EM Kcal", "Fibra%", "Grasa%", "Calcio%",
           "Fósforo%", "Lisina%", "Metionina%", "Colina mg/kg", "Precio/kg", "Categoría"]

DB_KEYS = ["nombre", "proteina", "em_kcal", "fibra", "grasa", "calcio",
           "fosforo", "lisina", "metionina", "colina_mgr", "precio_kg", "categoria"]

class TabIngredientes(QWidget):
    def __init__(self):
        super().__init__()
        self.insumos = []
        self.setup_ui()
        self.load_data()

    # ================================================================
    # UI
    # ================================================================
    def setup_ui(self):
        layout = QVBoxLayout(self)

        # ── Barra de controles superior ──────────────────────────────
        top = QHBoxLayout()

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍 Buscar insumo...")
        self.search_input.textChanged.connect(self.filter_table)

        self.combo_categoria = QComboBox()
        self.combo_categoria.addItems([
            "Todas", "Cereales", "Proteínas", "Grasas",
            "Subproductos", "Minerales", "Fibras", "Aditivos", "General"
        ])
        self.combo_categoria.currentTextChanged.connect(self.filter_table)

        self.btn_agregar  = QPushButton("➕ Agregar")
        self.btn_editar   = QPushButton("✏️ Editar")
        self.btn_eliminar = QPushButton("🗑️ Eliminar")
        self.btn_importar = QPushButton("📥 Importar Excel")
        self.btn_exportar = QPushButton("📤 Exportar Excel")

        # Estilos
        self.btn_agregar.setStyleSheet(
            "background-color: #2D7D46; color: white; font-weight: bold; padding: 5px 10px; border-radius: 4px;")
        self.btn_editar.setStyleSheet(
            "background-color: #2471A3; color: white; font-weight: bold; padding: 5px 10px; border-radius: 4px;")
        self.btn_eliminar.setStyleSheet(
            "background-color: #C0392B; color: white; font-weight: bold; padding: 5px 10px; border-radius: 4px;")
        self.btn_importar.setStyleSheet(
            "background-color: #7D3C98; color: white; font-weight: bold; padding: 5px 10px; border-radius: 4px;")
        self.btn_exportar.setStyleSheet(
            "background-color: #1E8449; color: white; font-weight: bold; padding: 5px 10px; border-radius: 4px;")

        # Conectar señales
        self.btn_agregar.clicked.connect(self.agregar_insumo)
        self.btn_editar.clicked.connect(self.editar_insumo)
        self.btn_eliminar.clicked.connect(self.eliminar_insumo)
        self.btn_importar.clicked.connect(self.importar_excel)
        self.btn_exportar.clicked.connect(self.exportar_excel)

        top.addWidget(self.search_input, 2)
        top.addWidget(self.combo_categoria)
        top.addWidget(self.btn_agregar)
        top.addWidget(self.btn_editar)
        top.addWidget(self.btn_eliminar)
        top.addWidget(self.btn_importar)
        top.addWidget(self.btn_exportar)

        layout.addLayout(top)

        # ── Tabla de insumos ─────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setColumnCount(len(HEADERS))
        self.table.setHorizontalHeaderLabels(HEADERS)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.doubleClicked.connect(self.editar_insumo)  # doble clic → editar

        layout.addWidget(self.table)

        # ── Contador de registros ────────────────────────────────────
        self.lbl_count = QLineEdit()
        self.lbl_count.setReadOnly(True)
        self.lbl_count.setStyleSheet("color: #888; border: none; background: transparent;")
        layout.addWidget(self.lbl_count)

    # ================================================================
    # Carga y actualización de tabla
    # ================================================================
    def load_data(self):
        """Recarga los insumos desde la base de datos y actualiza la tabla."""
        self.insumos = get_all_insumos()
        self.filter_table()

    def update_table(self, data):
        """Pinta la tabla con la lista de insumos proporcionada."""
        self.table.setRowCount(0)
        for row_data in data:
            row = self.table.rowCount()
            self.table.insertRow(row)
            for col, key in enumerate(DB_KEYS):
                val = row_data.get(key, "")
                if isinstance(val, float):
                    text = f"{val:.3f}" if key in ("calcio", "fosforo", "lisina", "metionina") else f"{val:.2f}"
                else:
                    text = str(val)
                item = QTableWidgetItem(text)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                # Guardar el ID oculto en el primer item
                if col == 0:
                    item.setData(Qt.ItemDataRole.UserRole, row_data.get('id'))
                self.table.setItem(row, col, item)
            

        self.lbl_count.setText(f"  {len(data)} insumo(s) mostrado(s)")

    def filter_table(self):
        """Filtra insumos por texto y categoría."""
        text = self.search_input.text().lower()
        cat  = self.combo_categoria.currentText()
        filtered = [
            ins for ins in self.insumos
            if text in ins['nombre'].lower()
            and (cat == "Todas" or cat == ins.get('categoria', ''))
        ]
        self.update_table(filtered)

    # ================================================================
    # Utilidad: obtener insumo seleccionado
    # ================================================================
    def _get_insumo_seleccionado(self):
        """Retorna el dict del insumo de la fila seleccionada, o None."""
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Sin selección", "Selecciona un insumo de la tabla primero.")
            return None, None
        id_insumo = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        # Buscar el insumo completo en self.insumos (incluye todos los campos)
        insumo = next((i for i in self.insumos if i.get('id') == id_insumo), None)
        return id_insumo, insumo

    # ================================================================
    # Acciones CRUD
    # ================================================================
    def agregar_insumo(self):
        dlg = InsumoDialog(self)
        if dlg.exec():
            data = dlg.get_data()
            # Asegura campos por defecto
            data.setdefault('proteina', 0)
            data.setdefault('em_kcal', 0)
            data.setdefault('fibra', 0)
            data.setdefault('grasa', 0)
            data.setdefault('calcio', 0)
            data.setdefault('fosforo', 0)
            data.setdefault('lisina', 0)
            data.setdefault('metionina', 0)
            data.setdefault('colina_mgr', 0)
            data.setdefault('precio_kg', 0)
            data.setdefault('categoria', 'General')
            if insert_insumo(data):
                self.load_data()
                QMessageBox.information(self, "✅ Insumo agregado",
                                        f"El insumo «{data['nombre']}» fue guardado correctamente.")
            else:
                QMessageBox.critical(self, "Error",
                                     "No se pudo guardar el insumo. Verifica que el nombre no esté duplicado.")

    def editar_insumo(self):
        id_insumo, insumo = self._get_insumo_seleccionado()
        if insumo is None:
            return
        dlg = InsumoDialog(self, insumo_data=insumo)
        if dlg.exec():
            data = dlg.get_data()
            if update_insumo(id_insumo, data):
                self.load_data()
                QMessageBox.information(self, "✅ Insumo actualizado",
                                        f"El insumo «{data['nombre']}» fue actualizado.")
            else:
                QMessageBox.critical(self, "Error", "No se pudo actualizar el insumo.")

    def eliminar_insumo(self):
        id_insumo, insumo = self._get_insumo_seleccionado()
        if insumo is None:
            return
        nombre = insumo.get('nombre', '')
        resp = QMessageBox.question(
            self, "Confirmar eliminación",
            f"¿Eliminar el insumo «{nombre}»?\nEsta acción no se puede deshacer.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if resp == QMessageBox.StandardButton.Yes:
            if delete_insumo(id_insumo):
                self.load_data()
                QMessageBox.information(self, "✅ Insumo eliminado",
                                        f"El insumo «{nombre}» fue eliminado.")
            else:
                QMessageBox.critical(self, "Error", "No se pudo eliminar el insumo.")

    # ================================================================
    # Importar / Exportar Excel
    # ================================================================
    def importar_excel(self):
        ruta, _ = QFileDialog.getOpenFileName(
            self, "Importar insumos desde Excel", "",
            "Archivos Excel (*.xlsx *.xls)"
        )
        if not ruta:
            return
        try:
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            ws = wb.active
            importados = 0
            errores = 0
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                try:
                    # Columnas esperadas: Nombre, Proteína, EM, Fibra, Grasa, Calcio, Fósforo, Lisina, Metionina, Colina, Precio, Categoría
                    nombre = str(row[0]).strip() if row[0] else None
                    if not nombre or nombre == "None":
                        continue
                    data = {
                        'nombre':     nombre,
                        'proteina':   float(row[1] or 0),
                        'em_kcal':    float(row[2] or 0),
                        'fibra':      float(row[3] or 0),
                        'grasa':      float(row[4] or 0),
                        'calcio':     float(row[5] or 0),
                        'fosforo':    float(row[6] or 0),
                        'lisina':     float(row[7] or 0),
                        'metionina':  float(row[8] or 0),
                        'colina_mgr': float(row[9] or 0),
                        'precio_kg':  float(row[10] or 0) if len(row) > 10 else 0,
                        'categoria':  str(row[11]).strip() if len(row) > 11 and row[11] else 'General',
                    }
                    if insert_insumo(data):
                        importados += 1
                    else:
                        errores += 1
                except Exception:
                    errores += 1
            self.load_data()
            QMessageBox.information(
                self, "✅ Importación completa",
                f"Se importaron {importados} insumo(s).\n"
                f"Omitidos/errores: {errores}.\n\n"
                "Formato esperado: Nombre | Proteína | EM Kcal | Fibra | Grasa | "
                "Calcio | Fósforo | Lisina | Metionina | Colina | Precio | Categoría"
            )
        except Exception as e:
            QMessageBox.critical(self, "Error al importar", str(e))

    def exportar_excel(self):
        ruta, _ = QFileDialog.getSaveFileName(
            self, "Exportar tabla de insumos", "tabla_insumos.xlsx",
            "Archivos Excel (*.xlsx)"
        )
        if not ruta:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Insumos"

            # Encabezados con estilo
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill("solid", fgColor="2D7D46")
            header_font = Font(color="FFFFFF", bold=True)
            ws.append(HEADERS)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Datos
            for ins in self.insumos:
                ws.append([
                    ins.get('nombre', ''),
                    ins.get('proteina', 0),
                    ins.get('em_kcal', 0),
                    ins.get('fibra', 0),
                    ins.get('grasa', 0),
                    ins.get('calcio', 0),
                    ins.get('fosforo', 0),
                    ins.get('lisina', 0),
                    ins.get('metionina', 0),
                    ins.get('colina_mgr', 0),
                    ins.get('precio_kg', 0),
                    ins.get('categoria', ''),
                ])

            # Ajustar ancho de columnas
            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col) + 3
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 30)

            wb.save(ruta)
            QMessageBox.information(self, "✅ Exportación completa",
                                    f"Tabla exportada en:\n{ruta}")
        except Exception as e:
            QMessageBox.critical(self, "Error al exportar", str(e))
